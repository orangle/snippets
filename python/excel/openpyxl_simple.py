from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"

for row in range(1, 10):
    ws1.append(range(10))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))


## merge cell
ws4 = wb.create_sheet(title="Merge")

alignment = Alignment(horizontal='center', vertical='center')
ws4.merge_cells(start_row=1, start_column=1, end_row=15, end_column=1)
country_cell = ws4.cell(row=1, column=1)
country_cell.value = "china"
country_cell.alignment = alignment

ws4.merge_cells(start_row=16, start_column=1, end_row=25, end_column=1)
c2 = ws4.cell(row=16, column=1)
c2.value = "usa"
c2.alignment = alignment

ws4.merge_cells(start_row=1, start_column=2, end_row=5, end_column=2)
isp_cell = ws4.cell(row=1, column=2)
isp_cell.value = 'china_mobile'

dist = ['10KB/s-', '10~100KB/s', '100~300KB/s', '300~500KB/s', '500KB/s+']

base = 1
for i in range(len(dist)):
    cell = ws4.cell(row=base + i, column=3)
    cell.value = dist[i]


wb.save(filename = dest_filename)